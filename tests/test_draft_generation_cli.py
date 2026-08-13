from __future__ import annotations

import hashlib
import json
import os
import shutil
import subprocess
import sys
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

import bank_config_compiler.cli as cli
from bank_config_compiler.draft_generation import (
    DraftGenerationError,
    DraftGenerationRequest,
    DraftProviderDiagnosticError,
    DraftProviderResult,
    ProviderCallMetadata,
    ProviderFailureCallEvidence,
    ProviderFailureEvidence,
    ProviderSubcallMetadata,
)


REPO_ROOT = Path(__file__).resolve().parents[1]


def bind_task(workspace: Path, *, task_id: str | None = None) -> None:
    raw_bytes = (workspace / "raw-doc.md").read_bytes()
    (workspace / "task.json").write_text(
        json.dumps(
            {
                "contractVersion": "phase0-task/v1",
                "taskId": task_id or workspace.name,
                "interfaceCode": "b2e0061",
                "messageFormat": "XML",
                "sourceDocument": "raw-doc.md",
                "sourceHash": "sha256:" + hashlib.sha256(raw_bytes).hexdigest(),
            },
            ensure_ascii=False,
            indent=2,
        )
        + "\n",
        encoding="utf-8",
        newline="",
    )


def run_cli(*args: str, cwd: Path = REPO_ROOT) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        [sys.executable, "-m", "bank_config_compiler", *args],
        cwd=cwd,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        check=False,
    )


def workbook_snapshot(path: Path) -> dict:
    workbook = load_workbook(path, data_only=False, keep_links=False)
    result = {"sheetnames": workbook.sheetnames, "sheets": {}}
    for sheet in workbook.worksheets:
        cells = []
        for row in sheet.iter_rows():
            for cell in row:
                value = (
                    "<generated-at>"
                    if sheet.title == "Overview" and cell.coordinate == "B4"
                    else cell.value
                )
                cells.append(
                    (
                        cell.coordinate,
                        value,
                        cell.data_type,
                        cell.number_format,
                        cell.fill.fill_type,
                        cell.fill.fgColor.rgb,
                        cell.font.bold,
                        cell.font.color.type if cell.font.color else None,
                        cell.font.color.rgb
                        if cell.font.color and cell.font.color.type == "rgb"
                        else None,
                        cell.alignment.vertical,
                        cell.alignment.wrap_text,
                    )
                )
        validations = sorted(
            (str(item.sqref), item.type, item.formula1, item.allow_blank)
            for item in sheet.data_validations.dataValidation
        )
        widths = sorted(
            (name, dimension.width)
            for name, dimension in sheet.column_dimensions.items()
            if dimension.width is not None
        )
        result["sheets"][sheet.title] = {
            "dimensions": sheet.dimensions,
            "freeze": sheet.freeze_panes,
            "filter": sheet.auto_filter.ref,
            "gridlines": sheet.sheet_view.showGridLines,
            "widths": widths,
            "validations": validations,
            "cells": cells,
        }
    workbook.close()
    return result


def prepare_docir_case(tmp_path: Path) -> tuple[Path, Path]:
    workspace = tmp_path / "workspace"
    workspace.mkdir()
    raw_bytes = (REPO_ROOT / "samples/golden/b2eboc-b2e0061/raw-doc.md").read_bytes()
    (workspace / "raw-doc.md").write_bytes(raw_bytes)
    bind_task(workspace)

    fixture_root = tmp_path / "fixture"
    fixture_root.mkdir()
    (fixture_root / "docir.md").write_bytes(
        (REPO_ROOT / "samples/golden/b2eboc-b2e0061/docir.expected.md").read_bytes()
    )
    (fixture_root / "notes.md").write_bytes(
        (REPO_ROOT / "samples/golden/b2eboc-b2e0061/review-notes.expected.md").read_bytes()
    )
    source_hash = f"sha256:{hashlib.sha256(raw_bytes).hexdigest()}"
    (fixture_root / "draft-stub-case.json").write_text(
        json.dumps(
            {
                "contractVersion": "draft-stub-case/v1",
                "caseId": "cli-docir-case",
                "responses": [
                    {
                        "request": {"artifactKind": "docir", "sourceHash": source_hash},
                        "artifactFile": "docir.md",
                        "reviewNotesFile": "notes.md",
                    }
                ],
            }
        ),
        encoding="utf-8",
        newline="",
    )
    return workspace, fixture_root


def test_generate_draft_docir_cli_writes_fixed_outputs(tmp_path: Path) -> None:
    workspace, fixture_root = prepare_docir_case(tmp_path)

    result = run_cli(
        "generate-draft",
        "docir",
        "--workspace",
        str(workspace),
        "--provider",
        "fixture",
        "--fixture-root",
        str(fixture_root),
    )

    assert result.returncode == 0, result.stderr
    assert "saved docir Draft" in result.stdout
    assert (workspace / "docir-draft.md").is_file()
    assert (workspace / "docir-review-notes.md").is_file()


def test_generate_draft_cli_requires_explicit_fixture_root(tmp_path: Path) -> None:
    workspace = tmp_path / "workspace"
    workspace.mkdir()
    (workspace / "raw-doc.md").write_text("# Raw\n", encoding="utf-8", newline="")
    bind_task(workspace)

    result = run_cli(
        "generate-draft",
        "docir",
        "--workspace",
        str(workspace),
        "--provider",
        "fixture",
    )

    assert result.returncode == 2
    assert "--fixture-root" in result.stderr


def test_generate_draft_cli_openai_chat_requires_runtime_api_key(
    tmp_path: Path,
    monkeypatch,
) -> None:
    monkeypatch.delenv("BANK_CONFIG_COMPILER_LLM_API_KEY", raising=False)
    workspace = tmp_path / "workspace"
    workspace.mkdir()
    (workspace / "raw-doc.md").write_text("# Raw bank document\n", encoding="utf-8")
    bind_task(workspace)

    result = run_cli(
        "generate-draft",
        "docir",
        "--workspace",
        str(workspace),
        "--provider",
        "openai-chat",
        "--chat-base-url",
        "https://example.invalid/v1",
        "--chat-model",
        "qwen-test-snapshot",
        "--attempt-id",
        "docir-001",
        cwd=tmp_path,
    )

    assert result.returncode == 2
    assert "BANK_CONFIG_COMPILER_LLM_API_KEY" in result.stderr
    assert not (workspace / "docir-draft.md").exists()


def test_main_loads_allowlisted_llm_configuration_from_cwd_dotenv(
    tmp_path: Path,
    monkeypatch,
) -> None:
    for name in (
        "BANK_CONFIG_COMPILER_LLM_API_KEY",
        "BANK_CONFIG_COMPILER_LLM_BASE_URL",
        "BANK_CONFIG_COMPILER_LLM_MODEL",
        "BANK_CONFIG_COMPILER_LLM_TIMEOUT_SECONDS",
        "UNRELATED_SETTING",
    ):
        monkeypatch.delenv(name, raising=False)
    (tmp_path / ".env").write_text(
        "BANK_CONFIG_COMPILER_LLM_API_KEY=dotenv-key\n"
        "BANK_CONFIG_COMPILER_LLM_BASE_URL=https://example.invalid/v1\n"
        "BANK_CONFIG_COMPILER_LLM_MODEL=qwen-test-snapshot\n"
        "BANK_CONFIG_COMPILER_LLM_TIMEOUT_SECONDS=45.5\n"
        "UNRELATED_SETTING=must-not-be-loaded\n",
        encoding="utf-8",
    )
    monkeypatch.chdir(tmp_path)

    def generate_draft(args) -> Path:
        assert cli._draft_provider(args) is provider
        return tmp_path / "docir-draft.md"

    provider = object()
    captured: dict[str, object] = {}

    def provider_factory(**kwargs: object) -> object:
        captured.update(kwargs)
        return provider

    monkeypatch.setattr(cli, "OpenAIChatDraftProvider", provider_factory)
    monkeypatch.setattr(cli, "_generate_draft", generate_draft)

    result = cli.main(
        [
            "generate-draft",
            "docir",
            "--workspace",
            str(tmp_path / "workspace"),
            "--provider",
            "openai-chat",
            "--attempt-id",
            "docir-001",
        ]
    )

    assert result == 0
    assert captured == {
        "api_key": "dotenv-key",
        "base_url": "https://example.invalid/v1",
        "model": "qwen-test-snapshot",
        "attempt_id": "docir-001",
        "timeout_seconds": 45.5,
        "docir_field_batch_size": 16,
    }
    assert "UNRELATED_SETTING" not in os.environ


def test_generate_draft_cli_returns_3_when_invalid_draft_was_published(
    tmp_path: Path,
    monkeypatch,
) -> None:
    output = tmp_path / "schemair-draft.json"
    monkeypatch.setattr(cli, "_generate_draft", lambda args: (output, 3))

    result = cli.main(
        [
            "generate-draft",
            "schemair",
            "--workspace",
            str(tmp_path / "workspace"),
            "--provider",
            "fixture",
            "--fixture-root",
            str(tmp_path / "fixture"),
            "--schema-id",
            "b2eboc-b2e0061-schema",
            "--schema-version",
            "v1",
        ]
    )

    assert result == 3


def test_process_environment_and_cli_override_dotenv_values(
    tmp_path: Path,
    monkeypatch,
) -> None:
    (tmp_path / ".env").write_text(
        "BANK_CONFIG_COMPILER_LLM_API_KEY=dotenv-key\n"
        "BANK_CONFIG_COMPILER_LLM_BASE_URL=https://dotenv.invalid/v1\n"
        "BANK_CONFIG_COMPILER_LLM_MODEL=dotenv-model\n"
        "BANK_CONFIG_COMPILER_LLM_TIMEOUT_SECONDS=10\n",
        encoding="utf-8",
    )
    monkeypatch.chdir(tmp_path)
    monkeypatch.setenv("BANK_CONFIG_COMPILER_LLM_API_KEY", "process-key")
    monkeypatch.setenv("BANK_CONFIG_COMPILER_LLM_MODEL", "process-model")
    runtime_environment = cli._load_runtime_environment()

    assert runtime_environment == {
        "BANK_CONFIG_COMPILER_LLM_API_KEY": "process-key",
        "BANK_CONFIG_COMPILER_LLM_BASE_URL": "https://dotenv.invalid/v1",
        "BANK_CONFIG_COMPILER_LLM_MODEL": "process-model",
        "BANK_CONFIG_COMPILER_LLM_TIMEOUT_SECONDS": "10",
    }


def test_openai_chat_cli_configuration_is_forwarded_without_secret_defaults(
    monkeypatch,
) -> None:
    captured: dict[str, object] = {}

    def provider_factory(**kwargs: object) -> object:
        captured.update(kwargs)
        return object()

    monkeypatch.setenv("BANK_CONFIG_COMPILER_LLM_API_KEY", "runtime-key")
    monkeypatch.setenv("BANK_CONFIG_COMPILER_LLM_BASE_URL", "https://env.invalid/v1")
    monkeypatch.setenv("BANK_CONFIG_COMPILER_LLM_MODEL", "env-model")
    monkeypatch.setenv("BANK_CONFIG_COMPILER_LLM_TIMEOUT_SECONDS", "10")
    monkeypatch.setattr(cli, "OpenAIChatDraftProvider", provider_factory)
    args = SimpleNamespace(
        draft_kind="docir",
        provider="openai-chat",
        fixture_root=None,
        chat_base_url="https://example.invalid/v1",
        chat_model="qwen-test-snapshot",
        chat_timeout_seconds=45.5,
        attempt_id="docir-001",
        docir_field_batch_size=8,
    )

    cli._draft_provider(args)

    assert captured == {
        "api_key": "runtime-key",
        "base_url": "https://example.invalid/v1",
        "model": "qwen-test-snapshot",
        "attempt_id": "docir-001",
        "timeout_seconds": 45.5,
        "docir_field_batch_size": 8,
    }


def test_docir_batch_size_must_be_positive() -> None:
    with pytest.raises(SystemExit):
        cli.build_parser().parse_args(
            [
                "generate-draft",
                "docir",
                "--workspace",
                "workspace",
                "--provider",
                "openai-chat",
                "--docir-field-batch-size",
                "0",
            ]
        )


def test_docir_batch_size_is_not_available_to_other_artifacts() -> None:
    with pytest.raises(SystemExit):
        cli.build_parser().parse_args(
            [
                "generate-draft",
                "schemair",
                "--workspace",
                "workspace",
                "--provider",
                "openai-chat",
                "--docir-field-batch-size",
                "8",
            ]
        )


def test_fixture_provider_rejects_explicit_docir_batch_size(tmp_path: Path) -> None:
    workspace, fixture_root = prepare_docir_case(tmp_path)

    result = run_cli(
        "generate-draft",
        "docir",
        "--workspace",
        str(workspace),
        "--provider",
        "fixture",
        "--fixture-root",
        str(fixture_root),
        "--docir-field-batch-size",
        "8",
    )

    assert result.returncode == 2
    assert "fixture provider does not accept chat configuration" in result.stderr


def test_dotenv_example_lists_only_supported_llm_configuration() -> None:
    example = (REPO_ROOT / ".env.example").read_text(encoding="utf-8")

    assert example == (
        "# Copy this file to .env and keep .env out of Git.\n"
        "BANK_CONFIG_COMPILER_LLM_API_KEY=\n"
        "BANK_CONFIG_COMPILER_LLM_BASE_URL=https://approved-provider.example/v1\n"
        "BANK_CONFIG_COMPILER_LLM_MODEL=approved-model-snapshot\n"
        "BANK_CONFIG_COMPILER_LLM_TIMEOUT_SECONDS=600\n"
    )


def test_generate_draft_cli_fails_closed_on_fixture_hash_mismatch(tmp_path: Path) -> None:
    workspace = tmp_path / "workspace"
    workspace.mkdir()
    (workspace / "raw-doc.md").write_text("different input", encoding="utf-8", newline="")
    bind_task(workspace)

    result = run_cli(
        "generate-draft", "docir", "--workspace", str(workspace),
        "--provider", "fixture",
        "--fixture-root", str(REPO_ROOT / "samples/draft-generation/b2eboc-b2e0061"),
    )

    assert result.returncode == 2
    assert "no exact response" in result.stderr
    assert not (workspace / "docir-draft.md").exists()
    assert not (workspace / "docir-review-notes.md").exists()


def test_generate_docir_failure_publishes_partial_provider_evidence(
    tmp_path: Path,
    monkeypatch,
) -> None:
    workspace = tmp_path / "phase0-docir-failure"
    workspace.mkdir()
    (workspace / "raw-doc.md").write_text(
        "# Raw bank document\n", encoding="utf-8", newline=""
    )
    bind_task(workspace)
    completed_responses = (
        '{"contractVersion":"docir-interface-envelope-segment/v2"}',
        '{"contractVersion":"docir-messages-outline-segment/v1"}',
    )
    partial_response = '{"contractVersion":"docir-field-details-segment/v2"'
    partial_hash = "sha256:" + hashlib.sha256(partial_response.encode()).hexdigest()

    class FailingProvider:
        name = "openai-chat"
        attempt_id = "docir-011"
        model = "qwen-test-snapshot"

        def generate(self, request, context):
            completed_calls = tuple(
                ProviderSubcallMetadata(
                    segment=segment,
                    outcome="succeeded",
                    response_complete=True,
                    response_content_hash=(
                        "sha256:" + hashlib.sha256(response.encode()).hexdigest()
                    ),
                    requested_model=self.model,
                    response_model=self.model,
                    response_id=f"chatcmpl-{sequence}",
                    started_at=f"2026-08-11T10:00:0{sequence}+08:00",
                    completed_at=f"2026-08-11T10:00:0{sequence + 1}+08:00",
                    finish_reason="stop",
                    prompt_contract_version="draft-prompt/v9",
                    segment_contract_version=contract,
                )
                for sequence, (segment, contract, response) in enumerate(
                    (
                        (
                            "interface-envelope",
                            "docir-interface-envelope-segment/v2",
                            completed_responses[0],
                        ),
                        (
                            "messages-outline",
                            "docir-messages-outline-segment/v1",
                            completed_responses[1],
                        ),
                    ),
                    start=1,
                )
            )
            failed_call = ProviderSubcallMetadata(
                segment="assembly-fields-001",
                outcome="failed",
                response_complete=False,
                response_content_hash=partial_hash,
                requested_model=self.model,
                response_model=self.model,
                response_id="chatcmpl-test",
                started_at="2026-08-11T10:00:00+08:00",
                completed_at="2026-08-11T10:01:00+08:00",
                prompt_contract_version="draft-prompt/v9",
                segment_contract_version="docir-field-details-segment/v2",
            )
            raise DraftProviderDiagnosticError(
                "chat stream failed: TimeoutError",
                evidence=ProviderFailureEvidence(
                    request=request,
                    metadata=ProviderCallMetadata(
                        provider_name=self.name,
                        attempt_id=self.attempt_id,
                        requested_model=self.model,
                        response_model=self.model,
                        response_id="chatcmpl-test",
                        started_at="2026-08-11T10:00:00+08:00",
                        completed_at="2026-08-11T10:01:00+08:00",
                        endpoint_fingerprint="sha256:" + "a" * 64,
                        prompt_contract_version="draft-prompt/v9",
                        calls=completed_calls + (failed_call,),
                        docir_field_batch_size=16,
                    ),
                    failure_stage="stream",
                    failure_detail="chat stream failed: TimeoutError",
                    error_type="TimeoutError",
                    response_complete=False,
                    response_text=partial_response,
                    finish_reason=None,
                    calls=tuple(
                        ProviderFailureCallEvidence(call, response)
                        for call, response in zip(
                            completed_calls,
                            completed_responses,
                            strict=True,
                        )
                    )
                    + (ProviderFailureCallEvidence(failed_call, partial_response),),
                    failed_segment="assembly-fields-001",
                ),
            )

    monkeypatch.setattr(cli, "_draft_provider", lambda args: FailingProvider())
    args = SimpleNamespace(
        workspace=workspace,
        draft_kind="docir",
        provider="openai-chat",
        overwrite=False,
    )

    with pytest.raises(DraftProviderDiagnosticError) as caught:
        cli._generate_draft(args)

    attempt = workspace / "provider-attempts" / "docir" / "docir-011"
    summary_path = attempt / "provider-failure-result.json"
    response_path = (
        attempt
        / "response-003-assembly-fields-001.txt"
    )
    summary = json.loads(summary_path.read_text(encoding="utf-8"))
    assert summary["contractVersion"] == "draft-provider-failure-result/v2"
    assert summary["artifactKind"] == "docir"
    assert summary["attemptId"] == "docir-011"
    assert summary["failureStage"] == "stream"
    assert summary["failureDetail"] == "chat stream failed: TimeoutError"
    assert summary["failedSegment"] == "assembly-fields-001"
    assert summary["responseComplete"] is False
    assert summary["responseContentHash"].startswith("sha256:")
    assert summary["endpointFingerprint"] == "sha256:" + "a" * 64
    serialized_summary = summary_path.read_text(encoding="utf-8")
    assert "test-key" not in serialized_summary
    assert "https://" not in serialized_summary
    assert "Raw bank document" not in serialized_summary
    assert [call["sequence"] for call in summary["calls"]] == [1, 2, 3]
    assert [call["segment"] for call in summary["calls"]] == [
        "interface-envelope",
        "messages-outline",
        "assembly-fields-001",
    ]
    assert [call["outcome"] for call in summary["calls"]] == [
        "succeeded",
        "succeeded",
        "failed",
    ]
    assert response_path.read_text(encoding="utf-8") == partial_response
    assert not (workspace / "docir-draft.md").exists()
    assert not (workspace / "docir-review-notes.md").exists()
    assert not (attempt / "provider-call-result.json").exists()
    expected_response_paths = {
        attempt / "response-001-interface-envelope.txt",
        attempt / "response-002-messages-outline.txt",
        response_path,
    }
    assert set(caught.value.failure_evidence_paths) == {
        summary_path,
        *expected_response_paths,
    }
    assert summary["responseContentHash"] in cli._safe_error(caught.value)


def test_generate_docir_request_failure_does_not_create_empty_response_file(
    tmp_path: Path,
    monkeypatch,
) -> None:
    workspace = tmp_path / "phase0-docir-request-failure"
    workspace.mkdir()
    (workspace / "raw-doc.md").write_text(
        "# Raw bank document\n", encoding="utf-8", newline=""
    )
    bind_task(workspace)

    class FailingProvider:
        name = "openai-chat"
        attempt_id = "docir-011"
        model = "qwen-test-snapshot"

        def generate(self, request, context):
            raise DraftProviderDiagnosticError(
                "chat request failed: TimeoutError",
                evidence=ProviderFailureEvidence(
                    request=request,
                    metadata=ProviderCallMetadata(
                        provider_name=self.name,
                        attempt_id=self.attempt_id,
                        requested_model=self.model,
                        started_at="2026-08-11T10:00:00+08:00",
                        completed_at="2026-08-11T10:00:01+08:00",
                        endpoint_fingerprint="sha256:" + "a" * 64,
                        prompt_contract_version="draft-prompt/v7",
                    ),
                    failure_stage="request",
                    failure_detail="chat request failed: TimeoutError",
                    error_type="TimeoutError",
                    response_complete=False,
                    response_text=None,
                    finish_reason=None,
                ),
            )

    monkeypatch.setattr(cli, "_draft_provider", lambda args: FailingProvider())
    args = SimpleNamespace(
        workspace=workspace,
        draft_kind="docir",
        provider="openai-chat",
        overwrite=False,
    )

    with pytest.raises(DraftProviderDiagnosticError) as caught:
        cli._generate_draft(args)

    attempt = workspace / "provider-attempts" / "docir" / "docir-011"
    summary_path = attempt / "provider-failure-result.json"
    summary = json.loads(summary_path.read_text(encoding="utf-8"))
    assert summary["failureStage"] == "request"
    assert summary["responseContentHash"] is None
    assert not list(attempt.glob("response-*.txt"))
    assert caught.value.failure_evidence_paths == (summary_path,)
    assert str(summary_path) in cli._safe_error(caught.value)


def test_schemair_provider_failure_consumes_attempt_before_retry(
    tmp_path: Path,
    monkeypatch,
) -> None:
    workspace = tmp_path / "phase0-schemair-request-failure"
    workspace.mkdir()
    (workspace / "raw-doc.md").write_text("# Raw\n", encoding="utf-8", newline="")
    bind_task(workspace)
    shutil.copyfile(
        REPO_ROOT / "samples/golden/b2eboc-b2e0061/docir.expected.md",
        workspace / "docir-final.md",
    )

    class FailingProvider:
        name = "openai-chat"
        attempt_id = "schemair-001"
        model = "qwen-test-snapshot"

        def __init__(self) -> None:
            self.calls = 0

        def generate(self, request, context):
            self.calls += 1
            raise DraftProviderDiagnosticError(
                "chat request failed: TimeoutError",
                evidence=ProviderFailureEvidence(
                    request=request,
                    metadata=ProviderCallMetadata(
                        provider_name=self.name,
                        attempt_id=self.attempt_id,
                        requested_model=self.model,
                        started_at="2026-08-12T10:00:00+08:00",
                        completed_at="2026-08-12T10:00:01+08:00",
                        endpoint_fingerprint="sha256:" + "a" * 64,
                        prompt_contract_version="draft-prompt/v9",
                    ),
                    failure_stage="request",
                    failure_detail="chat request failed: TimeoutError",
                    error_type="TimeoutError",
                    response_complete=False,
                    response_text=None,
                    finish_reason=None,
                ),
            )

    provider = FailingProvider()
    monkeypatch.setattr(cli, "_draft_provider", lambda args: provider)
    args = SimpleNamespace(
        workspace=workspace,
        draft_kind="schemair",
        provider="openai-chat",
        overwrite=False,
        schema_id="b2eboc-b2e0061-schema",
        schema_version="v1",
    )

    with pytest.raises(DraftProviderDiagnosticError):
        cli._generate_draft(args)

    attempt = workspace / "provider-attempts/schemair/schemair-001"
    summary = json.loads(
        (attempt / "provider-failure-result.json").read_text(encoding="utf-8")
    )
    assert summary["artifactKind"] == "schemair"
    assert summary["attemptId"] == "schemair-001"
    assert not (workspace / "schemair-draft.json").exists()

    with pytest.raises(DraftGenerationError, match="attempt ID.*already exists"):
        cli._generate_draft(args)
    assert provider.calls == 1


def test_schemair_materialization_failure_saves_candidate_and_consumes_attempt(
    tmp_path: Path,
    monkeypatch,
) -> None:
    workspace = tmp_path / "phase0-schemair-materialization-failure"
    workspace.mkdir()
    (workspace / "raw-doc.md").write_text("# Raw\n", encoding="utf-8", newline="")
    bind_task(workspace)
    shutil.copyfile(
        REPO_ROOT / "samples/golden/b2eboc-b2e0061/docir.expected.md",
        workspace / "docir-final.md",
    )
    schema = json.loads(
        (
            REPO_ROOT
            / "samples/trusted-chain/b2eboc-b2e0061/schemair-final.json"
        ).read_text(encoding="utf-8")
    )
    schema["envelope"]["fields"].pop()
    candidate_text = json.dumps(schema, ensure_ascii=False)

    class InvalidCandidateProvider:
        name = "openai-chat"
        attempt_id = "schemair-002"
        model = "qwen-test-snapshot"

        def __init__(self) -> None:
            self.calls = 0

        def generate(self, request, context):
            self.calls += 1
            return DraftProviderResult(
                response_text=json.dumps(
                    {
                        "contractVersion": "draft-provider-response/v1",
                        "artifactKind": "schemair",
                        "artifactContent": candidate_text,
                        "reviewNotes": "# Review\n\nPending.\n",
                    },
                    ensure_ascii=False,
                ),
                metadata=ProviderCallMetadata(
                    provider_name=self.name,
                    attempt_id=self.attempt_id,
                    requested_model=self.model,
                    started_at="2026-08-12T10:00:00+08:00",
                    completed_at="2026-08-12T10:00:01+08:00",
                    endpoint_fingerprint="sha256:" + "a" * 64,
                    prompt_contract_version="draft-prompt/v9",
                ),
            )

    provider = InvalidCandidateProvider()
    monkeypatch.setattr(cli, "_draft_provider", lambda args: provider)
    args = SimpleNamespace(
        workspace=workspace,
        draft_kind="schemair",
        provider="openai-chat",
        overwrite=False,
        schema_id="b2eboc-b2e0061-schema",
        schema_version="v1",
    )

    with pytest.raises(DraftProviderDiagnosticError, match="cannot be materialized"):
        cli._generate_draft(args)

    attempt = workspace / "provider-attempts/schemair/schemair-002"
    summary = json.loads(
        (attempt / "provider-failure-result.json").read_text(encoding="utf-8")
    )
    assert summary["failureStage"] == "materialization"
    assert summary["candidateContentHash"].startswith("sha256:")
    assert summary["calls"][0]["outcome"] == "succeeded"
    assert (attempt / "candidate.json").read_text(encoding="utf-8") == candidate_text
    assert not (workspace / "schemair-draft.json").exists()

    with pytest.raises(DraftGenerationError, match="attempt ID.*already exists"):
        cli._generate_draft(args)
    assert provider.calls == 1


def test_provider_failure_attempt_cannot_be_overwritten(tmp_path: Path) -> None:
    workspace = tmp_path / "phase0-docir-overwrite-failure"
    workspace.mkdir()
    (workspace / "raw-doc.md").write_text("# Raw\n", encoding="utf-8", newline="")
    bind_task(workspace)

    def diagnostic(response_text: str | None) -> DraftProviderDiagnosticError:
        request = DraftGenerationRequest(
            task_id="phase0-docir-overwrite-failure",
            artifact_kind="docir",
            source_hash="sha256:" + "1" * 64,
        )
        return DraftProviderDiagnosticError(
            "provider failed",
            evidence=ProviderFailureEvidence(
                request=request,
                metadata=ProviderCallMetadata(
                    provider_name="openai-chat",
                    attempt_id="docir-011",
                    requested_model="qwen-test-snapshot",
                    started_at="2026-08-11T10:00:00+08:00",
                    completed_at="2026-08-11T10:00:01+08:00",
                    endpoint_fingerprint="sha256:" + "a" * 64,
                    prompt_contract_version="draft-prompt/v7",
                ),
                failure_stage="stream" if response_text is not None else "request",
                failure_detail="provider failed",
                error_type="TimeoutError",
                response_complete=False,
                response_text=response_text,
                finish_reason=None,
            ),
        )

    cli.publish_provider_failure(workspace, diagnostic("partial response"))
    response_path = (
        workspace
        / "provider-attempts"
        / "docir"
        / "docir-011"
        / "response-001-complete-artifact.txt"
    )
    assert response_path.is_file()

    with pytest.raises(DraftGenerationError, match="attempt ID.*already exists"):
        cli.publish_provider_failure(workspace, diagnostic(None), overwrite=True)

    assert response_path.read_text(encoding="utf-8") == "partial response"


def test_generate_draft_cli_never_promotes_docir_draft(tmp_path: Path) -> None:
    workspace = tmp_path / "workspace"
    workspace.mkdir()
    fixture_root = REPO_ROOT / "samples/draft-generation/b2eboc-b2e0061"
    shutil.copyfile(
        REPO_ROOT / "samples/golden/b2eboc-b2e0061/raw-doc.md",
        workspace / "raw-doc.md",
    )
    bind_task(workspace)

    docir = run_cli(
        "generate-draft", "docir", "--workspace", str(workspace),
        "--provider", "fixture", "--fixture-root", str(fixture_root),
    )
    assert docir.returncode == 0, docir.stderr
    assert (workspace / "docir-draft.md").is_file()
    assert not (workspace / "docir-final.md").exists()

    schemair = run_cli(
        "generate-draft", "schemair", "--workspace", str(workspace),
        "--provider", "fixture", "--fixture-root", str(fixture_root),
        "--schema-id", "b2eboc-b2e0061-schema", "--schema-version", "v1",
    )
    assert schemair.returncode == 2
    assert "docir-final.md is missing" in schemair.stderr
    assert not (workspace / "schemair-draft.json").exists()


def test_generate_draft_cli_completes_controlled_b2e0061_workflow(tmp_path: Path) -> None:
    workspace = tmp_path / "workspace"
    workspace.mkdir()
    fixture_root = REPO_ROOT / "samples/draft-generation/b2eboc-b2e0061"
    raw_source = REPO_ROOT / "samples/golden/b2eboc-b2e0061/raw-doc.md"
    shutil.copyfile(raw_source, workspace / "raw-doc.md")
    bind_task(workspace)

    docir = run_cli(
        "generate-draft", "docir", "--workspace", str(workspace),
        "--provider", "fixture", "--fixture-root", str(fixture_root),
    )
    assert docir.returncode == 0, docir.stderr

    # Human gate 只能通过显式装载已审核 fixture 表达；严禁把本次生成的 Draft 自动提升为 Final。
    shutil.copyfile(fixture_root / "docir-final.md", workspace / "docir-final.md")
    assert (workspace / "docir-final.md").read_bytes() == (
        fixture_root / "docir-final.md"
    ).read_bytes()
    schemair = run_cli(
        "generate-draft", "schemair", "--workspace", str(workspace),
        "--provider", "fixture", "--fixture-root", str(fixture_root),
        "--schema-id", "b2eboc-b2e0061-schema", "--schema-version", "v1",
    )
    assert schemair.returncode == 0, schemair.stderr

    trusted = REPO_ROOT / "samples/trusted-chain/b2eboc-b2e0061"
    for name in ("schemair-final.json", "schemair-validation-result.json"):
        shutil.copyfile(trusted / name, workspace / name)
    for direction, template_id in (
        ("assembly", "b2e0061-assembly-common"),
        ("parse", "b2e0061-parse-common"),
    ):
        standard = run_cli(
            "generate-draft", "standard", "--workspace", str(workspace),
            "--provider", "fixture", "--fixture-root", str(fixture_root),
            "--direction", direction, "--standard-version", "v1",
            "--standard-id", f"b2e0061-{direction}-standard",
            "--rule-package", str(REPO_ROOT / "configuration-rules/v1"),
        )
        assert standard.returncode == 0, standard.stderr

        standard_dir = workspace / "standards" / direction / "v1"
        trusted_standard_dir = trusted / "standards" / direction / "v1"
        for name in ("standard-final.json", "standard-validation-result.json"):
            shutil.copyfile(trusted_standard_dir / name, standard_dir / name)
        template = run_cli(
            "generate-draft", "template", "--workspace", str(workspace),
            "--provider", "fixture", "--fixture-root", str(fixture_root),
            "--direction", direction, "--standard-version", "v1",
            "--template-id", template_id, "--template-version", "v1",
            "--rule-package", str(REPO_ROOT / "configuration-rules/v2"),
        )
        assert template.returncode == 0, template.stderr
        template_dir = workspace / "templates" / direction / template_id / "v1"
        assert (template_dir / "template-draft.json").is_file()

        trusted_template_dir = trusted / "templates" / direction / "v1"
        for name in ("template-final.json", "template-validation-result.json"):
            shutil.copyfile(trusted_template_dir / name, template_dir / name)

        phase0_args = [
            "--workspace", str(workspace),
            "--direction", direction,
            "--standard-version", "v1",
            "--template-id", template_id,
            "--template-version", "v1",
            "--standard-rule-package", str(REPO_ROOT / "configuration-rules/v1"),
            "--template-rule-package", str(REPO_ROOT / "configuration-rules/v2"),
        ]
        checked = run_cli("check", "--profile", "phase0", *phase0_args)
        assert checked.returncode == 0, checked.stderr

        workbook = run_cli(
            "generate-workbook",
            *phase0_args,
            "--standard-action",
            "CREATE",
        )
        assert workbook.returncode == 0, workbook.stderr
        actual_workbook = template_dir / "configuration-workbook.xlsx"
        expected_workbook = trusted_template_dir / "configuration-workbook.xlsx"
        assert workbook_snapshot(actual_workbook) == workbook_snapshot(expected_workbook)
