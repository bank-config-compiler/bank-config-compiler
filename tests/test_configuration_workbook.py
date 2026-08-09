from __future__ import annotations

import json
import logging
from copy import deepcopy
from datetime import datetime, timezone
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

import bank_config_compiler.configuration_workbook as workbook_module
from bank_config_compiler.configuration_rules import load_rule_package
from bank_config_compiler.configuration_workbook import (
    WORKBOOK_FORMAT_VERSION,
    WorkbookGenerationError,
    _add_template_sheet,
    _expression_rows,
    _set_cell,
    _warning_rows,
    generate_configuration_workbook,
    validate_configuration_workbook_inputs,
)
from bank_config_compiler.interface_template_validator import validate_interface_template


REPO_ROOT = Path(__file__).parents[1]
CHAIN_ROOT = REPO_ROOT / "samples" / "trusted-chain" / "b2eboc-b2e0061"


def _json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def _assembly_inputs() -> dict:
    return _direction_inputs("assembly")


def _direction_inputs(direction: str) -> dict:
    template_dir = CHAIN_ROOT / "templates" / direction / "v1"
    standard_dir = CHAIN_ROOT / "standards" / direction / "v1"
    return {
        "schemair": _json(CHAIN_ROOT / "schemair-final.json"),
        "schemair_validation_result": _json(CHAIN_ROOT / "schemair-validation-result.json"),
        "standard": _json(standard_dir / "standard-final.json"),
        "standard_validation_result": _json(standard_dir / "standard-validation-result.json"),
        "template": _json(template_dir / "template-final.json"),
        "template_validation_result": _json(template_dir / "template-validation-result.json"),
        "standard_rule_package": load_rule_package(REPO_ROOT / "configuration-rules" / "v1"),
        "template_rule_package": load_rule_package(REPO_ROOT / "configuration-rules" / "v2"),
    }


def test_final_chain_generates_fixed_workbook_structure(tmp_path: Path) -> None:
    inputs = _assembly_inputs()

    validate_configuration_workbook_inputs(**inputs)
    output = generate_configuration_workbook(
        **inputs,
        standard_action="CREATE",
        output_path=tmp_path / "configuration-workbook.xlsx",
        generated_at=datetime(2026, 8, 9, 12, 0, tzinfo=timezone.utc),
    )

    assert WORKBOOK_FORMAT_VERSION == "v1"
    assert output.is_file()
    workbook = load_workbook(output, data_only=False)
    assert workbook.sheetnames == [
        "Overview",
        "Interface Standard",
        "Interface Template",
        "Value Expressions",
        "Warnings",
        "Rule References",
        "Legend",
    ]
    assert workbook["Interface Standard"].max_row == 37
    assert workbook["Interface Template"].max_row == 27
    assert workbook["Value Expressions"].max_row == 31
    assert workbook["Warnings"].max_row == 39


def _issue_codes(error: WorkbookGenerationError) -> set[str]:
    return {item["code"] for item in error.issues}


def test_parse_chain_projects_expected_rows_and_warnings(tmp_path: Path) -> None:
    output = generate_configuration_workbook(
        **_direction_inputs("parse"),
        standard_action="CREATE",
        output_path=tmp_path / "parse.xlsx",
        generated_at=datetime(2026, 8, 9, 12, 0, tzinfo=timezone.utc),
    )

    workbook = load_workbook(output, data_only=False)
    assert workbook["Interface Standard"].max_row == 20
    assert workbook["Interface Template"].max_row == 9
    assert workbook["Value Expressions"].max_row == 14
    assert workbook["Warnings"].max_row == 16

    warning_headers = [cell.value for cell in workbook["Warnings"][1]]
    categories = [
        row[warning_headers.index("Category")].value
        for row in workbook["Warnings"].iter_rows(min_row=2)
    ]
    assert categories.count("SCHEMA_STANDARD_DIFFERENCE") == 4
    assert categories.count("VALIDATOR") == 11


def test_warnings_project_every_condition_for_same_standard_field() -> None:
    inputs = _assembly_inputs()
    standard = deepcopy(inputs["standard"])
    target = next(field for field in standard["fields"] if field.get("conditionalConstraints"))
    second_condition = deepcopy(target["conditionalConstraints"][0])
    second_condition["conditionId"] = "b2e0061-assembly-condition-2"
    second_condition["schemaIrConditionIndex"] = 1
    second_condition["literal"] = "3"
    second_condition["sourceText"] = "第二条已确认银行条件。"
    target["conditionalConstraints"].append(second_condition)

    rows = _warning_rows(
        inputs["schemair_validation_result"],
        standard,
        inputs["standard_validation_result"],
        inputs["template"],
        inputs["template_validation_result"],
    )

    condition_rows = [
        row
        for row in rows
        if row[2] == target["fieldId"] and row[3] == "BANK_CONDITIONAL_CONSTRAINT"
    ]
    assert len(condition_rows) == 2
    assert {
        row[4].split(" EQUALS ", 1)[1].split(" =>", 1)[0]
        for row in condition_rows
    } == {"2", "3"}


def test_generator_rejects_forged_validation_result(tmp_path: Path) -> None:
    inputs = _assembly_inputs()
    inputs["template_validation_result"] = deepcopy(inputs["template_validation_result"])
    inputs["template_validation_result"]["summary"]["warningCount"] = 0

    with pytest.raises(WorkbookGenerationError) as captured:
        generate_configuration_workbook(
            **inputs,
            standard_action="CREATE",
            output_path=tmp_path / "forged.xlsx",
            generated_at=datetime(2026, 8, 9, 12, 0, tzinfo=timezone.utc),
        )

    assert _issue_codes(captured.value) == {"VALIDATION_RESULT_MISMATCH"}
    assert not (tmp_path / "forged.xlsx").exists()


def test_generator_requires_two_exact_released_rule_packages(tmp_path: Path) -> None:
    inputs = _assembly_inputs()
    inputs["standard_rule_package"] = inputs["template_rule_package"]

    with pytest.raises(WorkbookGenerationError) as captured:
        generate_configuration_workbook(
            **inputs,
            standard_action="CREATE",
            output_path=tmp_path / "wrong-rules.xlsx",
            generated_at=datetime(2026, 8, 9, 12, 0, tzinfo=timezone.utc),
        )

    assert _issue_codes(captured.value) == {"RULE_PACKAGE_VERSION_MISMATCH"}


def test_generator_rejects_final_template_with_redacted_mapping_reference(tmp_path: Path) -> None:
    inputs = _assembly_inputs()
    template = deepcopy(inputs["template"])
    target = next(config for config in template["fieldConfigs"] if config["bindingKind"] == "VALUE")
    target["valueExpression"] = {
        "mode": "MAPPING",
        "sequence": 1,
        "ruleReferences": ["TPL.VALUE.MAPPING"],
        "assemblyFieldRef": "chargeBearer",
        "mappingRuleName": "Swift-CompanyName-List",
    }
    inputs["template"] = template
    inputs["template_validation_result"] = validate_interface_template(
        template,
        standard=inputs["standard"],
        rule_package=inputs["template_rule_package"],
    )

    with pytest.raises(WorkbookGenerationError) as captured:
        generate_configuration_workbook(
            **inputs,
            standard_action="CREATE",
            output_path=tmp_path / "redacted-mapping.xlsx",
            generated_at=datetime(2026, 8, 9, 12, 0, tzinfo=timezone.utc),
        )

    assert "ARTIFACT_NOT_FINAL_ELIGIBLE" in _issue_codes(captured.value)
    assert not (tmp_path / "redacted-mapping.xlsx").exists()


@pytest.mark.parametrize("action", ["create", "", True, None])
def test_generator_rejects_implicit_or_invalid_standard_action(tmp_path: Path, action: object) -> None:
    with pytest.raises(WorkbookGenerationError) as captured:
        generate_configuration_workbook(
            **_assembly_inputs(),
            standard_action=action,  # type: ignore[arg-type]
            output_path=tmp_path / "invalid-action.xlsx",
            generated_at=datetime(2026, 8, 9, 12, 0, tzinfo=timezone.utc),
        )

    assert _issue_codes(captured.value) == {"INVALID_STANDARD_ACTION"}


def test_generator_requires_aware_datetime_and_explicit_overwrite(tmp_path: Path) -> None:
    output = tmp_path / "configuration-workbook.xlsx"
    inputs = _assembly_inputs()

    with pytest.raises(WorkbookGenerationError) as captured:
        generate_configuration_workbook(
            **inputs,
            standard_action="CREATE",
            output_path=output,
            generated_at=datetime(2026, 8, 9, 12, 0),
        )
    assert _issue_codes(captured.value) == {"INVALID_GENERATED_AT"}

    generate_configuration_workbook(
        **inputs,
        standard_action="CREATE",
        output_path=output,
        generated_at=datetime(2026, 8, 9, 12, 0, tzinfo=timezone.utc),
    )
    original = output.read_bytes()
    with pytest.raises(WorkbookGenerationError) as captured:
        generate_configuration_workbook(
            **inputs,
            standard_action="UPDATE",
            output_path=output,
            generated_at=datetime(2026, 8, 9, 12, 1, tzinfo=timezone.utc),
        )
    assert _issue_codes(captured.value) == {"OUTPUT_ALREADY_EXISTS"}
    assert output.read_bytes() == original

    generate_configuration_workbook(
        **inputs,
        standard_action="UPDATE",
        output_path=output,
        generated_at=datetime(2026, 8, 9, 12, 1, tzinfo=timezone.utc),
        overwrite=True,
    )
    workbook = load_workbook(output)
    overview = {row[0].value: row[1].value for row in workbook["Overview"].iter_rows(min_row=2)}
    assert overview["Standard Action"] == "UPDATE"
    assert "人工" in overview["Update Notice"]


def test_reuse_standard_rows_are_read_only_without_data_validation(tmp_path: Path) -> None:
    output = generate_configuration_workbook(
        **_assembly_inputs(),
        standard_action="REUSE",
        output_path=tmp_path / "reuse.xlsx",
        generated_at=datetime(2026, 8, 9, 12, 0, tzinfo=timezone.utc),
    )
    workbook = load_workbook(output)
    sheet = workbook["Interface Standard"]
    headers = [cell.value for cell in sheet[1]]
    execution = headers.index("Execution Status") + 1
    verification = headers.index("Verification Status") + 1

    assert {sheet.cell(row=row, column=execution).value for row in range(2, sheet.max_row + 1)} == {
        "NOT_APPLICABLE"
    }
    assert {sheet.cell(row=row, column=verification).value for row in range(2, sheet.max_row + 1)} == {
        "NOT_APPLICABLE"
    }
    assert len(sheet.data_validations.dataValidation) == 0


def test_all_business_strings_are_literal_cells_not_formulas(tmp_path: Path) -> None:
    workbook = Workbook()
    cell = workbook.active["A1"]

    _set_cell(cell, "=1+1")

    assert cell.value == "=1+1"
    assert cell.data_type == "s"

    for unsafe in ("contains\x01control", "x" * 32768, "value <REDACTED>"):
        with pytest.raises(WorkbookGenerationError):
            _set_cell(cell, unsafe)


def test_mapping_and_replacement_projection_exposes_names_not_catalog_entries() -> None:
    controlled = _json(REPO_ROOT / "tests" / "fixtures" / "interface-template-v1" / "mapping-replacement.json")
    expression = controlled["mappingExpression"]

    expression_rows = _expression_rows(
        "controlled-template",
        "controlled-target",
        "FIELD_VALUE",
        None,
        expression,
    )
    rendered_expression = json.dumps(expression_rows, ensure_ascii=False)
    assert "BDC-ChargeBearer-List" in rendered_expression
    assert "chargeBearer" in rendered_expression
    assert "DEBT" not in rendered_expression
    assert "OUR" not in rendered_expression

    config = {
        "bindingKind": "VALUE",
        "parseTarget": {
            "parseFieldRef": "controlled-target",
            "name": "controlledTarget",
            "parentPath": "Root",
            "fullPath": "Root.controlledTarget",
            "dataType": "STRING",
        },
        "valueExpression": expression,
        "processingPolicies": {
            "emptyHandling": "BLANK",
            "overlengthHandling": "INTERCEPT",
            "rowLimit": 1,
            "chineseCharacterLength": "STANDARD_1",
            "replacementRuleName": controlled["replacementRuleName"],
        },
        "ruleReferences": ["TPL.VALUE.MAPPING"],
        "confidence": 1.0,
        "uncertain": False,
    }
    standard = {
        "fields": [
            {
                "fieldId": "chargeBearer",
                "fieldName": "chargeBearer",
                "parentPath": "Root",
                "fullPath": "Root.chargeBearer",
                "required": False,
                "lengthLimit": {"state": "NO_CONSTRAINT"},
                "dataType": "String",
            }
        ]
    }
    template = {"direction": "PARSE", "fieldConfigs": [config]}
    workbook = Workbook()
    workbook.remove(workbook.active)
    _add_template_sheet(workbook, standard, template, {"issues": []})
    values = [cell.value for cell in workbook["Interface Template"][2]]
    rendered_row = json.dumps(values, ensure_ascii=False)
    assert "Swift_illegalCharacter_List_For_ING_Turkey" in rendered_row
    assert "DEBT" not in values
    assert "OUR" not in values


def test_formula_scan_and_status_validations_are_present(tmp_path: Path) -> None:
    output = generate_configuration_workbook(
        **_assembly_inputs(),
        standard_action="CREATE",
        output_path=tmp_path / "formula-scan.xlsx",
        generated_at=datetime(2026, 8, 9, 12, 0, tzinfo=timezone.utc),
    )
    workbook = load_workbook(output, data_only=False)

    assert all(cell.data_type != "f" for sheet in workbook.worksheets for row in sheet.iter_rows() for cell in row)
    assert len(workbook["Interface Standard"].data_validations.dataValidation) == 2
    assert len(workbook["Interface Template"].data_validations.dataValidation) == 2


def test_generation_logs_only_safe_metadata(tmp_path: Path, caplog: pytest.LogCaptureFixture) -> None:
    output = tmp_path / "logged.xlsx"
    inputs = _assembly_inputs()
    logger_name = "bank_config_compiler.configuration_workbook"

    with caplog.at_level(logging.DEBUG, logger=logger_name):
        generate_configuration_workbook(
            **inputs,
            standard_action="CREATE",
            output_path=output,
            generated_at=datetime(2026, 8, 9, 12, 0, tzinfo=timezone.utc),
        )

    records = [record for record in caplog.records if record.name == logger_name]
    assert [record.outcome for record in records] == ["started", "succeeded"]
    assert records[-1].interface_code == "b2e0061"
    assert records[-1].direction == "ASSEMBLY"
    assert records[-1].template_id == "b2e0061-assembly-common"
    standard_keys = set(logging.makeLogRecord({}).__dict__)
    allowed_metadata = {
        "component",
        "interface_code",
        "direction",
        "template_id",
        "outcome",
        "sheet_count",
        "row_counts",
    }
    for record in records:
        custom_keys = set(record.__dict__) - standard_keys - {"message", "asctime"}
        assert custom_keys <= allowed_metadata
    rendered = repr([record.__dict__ for record in records])
    assert "bocb2e.assembly.termid" not in rendered
    assert "最大9223372036854775807" not in rendered


def test_atomic_publish_failure_removes_temporary_workbook(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    output = tmp_path / "atomic-failure.xlsx"

    def fail_publish(_source: Path, _target: Path) -> None:
        raise OSError("simulated publish failure")

    monkeypatch.setattr(workbook_module.os, "link", fail_publish)

    with pytest.raises(WorkbookGenerationError) as captured:
        generate_configuration_workbook(
            **_assembly_inputs(),
            standard_action="CREATE",
            output_path=output,
            generated_at=datetime(2026, 8, 9, 12, 0, tzinfo=timezone.utc),
        )

    assert _issue_codes(captured.value) == {"WORKBOOK_WRITE_FAILED"}
    assert not output.exists()
    assert list(tmp_path.iterdir()) == []
