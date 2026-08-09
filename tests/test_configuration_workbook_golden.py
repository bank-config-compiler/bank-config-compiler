from __future__ import annotations

import json
import re
import zipfile
from datetime import datetime, timedelta, timezone
from pathlib import Path

import pytest
from openpyxl import load_workbook

from bank_config_compiler.configuration_rules import load_rule_package
from bank_config_compiler.configuration_workbook import generate_configuration_workbook


REPO_ROOT = Path(__file__).parents[1]
CHAIN_ROOT = REPO_ROOT / "samples" / "trusted-chain" / "b2eboc-b2e0061"
GOLDEN_GENERATED_AT = datetime(2026, 8, 9, 20, 0, tzinfo=timezone(timedelta(hours=8)))


def _json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def _inputs(direction: str) -> dict:
    standard_dir = CHAIN_ROOT / "standards" / direction / "v1"
    template_dir = CHAIN_ROOT / "templates" / direction / "v1"
    return {
        "schemair": _json(CHAIN_ROOT / "schemair-final.json"),
        "schemair_validation_result": _json(CHAIN_ROOT / "schemair-validation-result.json"),
        "standard": _json(standard_dir / "standard-final.json"),
        "standard_validation_result": _json(standard_dir / "standard-validation-result.json"),
        "template": _json(template_dir / "template-final.json"),
        "template_validation_result": _json(template_dir / "template-validation-result.json"),
        "standard_rule_package": load_rule_package(REPO_ROOT / "configuration-rules" / "v1"),
        "template_rule_package": load_rule_package(REPO_ROOT / "configuration-rules" / "v2"),
    }


def _workbook_snapshot(path: Path) -> dict:
    workbook = load_workbook(path, data_only=False, keep_links=False)
    result = {"sheetnames": workbook.sheetnames, "sheets": {}}
    for sheet in workbook.worksheets:
        cells = []
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if sheet.title == "Overview" and cell.column == 2 and sheet.cell(cell.row, 1).value == "Generated At":
                    value = "<generated-at>"
                cells.append(
                    (
                        cell.coordinate,
                        value,
                        cell.data_type,
                        cell.number_format,
                        cell.fill.fill_type,
                        cell.fill.fgColor.rgb,
                        cell.font.bold,
                        cell.font.color.type if cell.font.color else None,
                        cell.font.color.rgb if cell.font.color and cell.font.color.type == "rgb" else None,
                        cell.alignment.vertical,
                        cell.alignment.wrap_text,
                    )
                )
        validations = sorted(
            (
                str(item.sqref),
                item.type,
                item.formula1,
                item.allow_blank,
            )
            for item in sheet.data_validations.dataValidation
        )
        widths = sorted(
            (name, dimension.width)
            for name, dimension in sheet.column_dimensions.items()
            if dimension.width is not None
        )
        result["sheets"][sheet.title] = {
            "dimensions": sheet.dimensions,
            "freeze": sheet.freeze_panes,
            "filter": sheet.auto_filter.ref,
            "gridlines": sheet.sheet_view.showGridLines,
            "widths": widths,
            "validations": validations,
            "cells": cells,
        }
    workbook.close()
    return result


@pytest.mark.parametrize(
    ("direction", "expected_counts"),
    [
        ("assembly", (36, 26, 30, 38)),
        ("parse", (19, 8, 13, 15)),
    ],
)
def test_committed_workbook_matches_fresh_structured_generation(
    tmp_path: Path,
    direction: str,
    expected_counts: tuple[int, int, int, int],
) -> None:
    expected = CHAIN_ROOT / "templates" / direction / "v1" / "configuration-workbook.xlsx"
    assert expected.is_file()
    actual = tmp_path / f"{direction}.xlsx"

    generate_configuration_workbook(
        **_inputs(direction),
        standard_action="CREATE",
        output_path=actual,
        generated_at=GOLDEN_GENERATED_AT,
    )

    assert _workbook_snapshot(actual) == _workbook_snapshot(expected)
    workbook = load_workbook(expected, data_only=False, keep_links=False)
    standard_count, template_count, expression_count, warning_count = expected_counts
    assert workbook["Interface Standard"].max_row - 1 == standard_count
    assert workbook["Interface Template"].max_row - 1 == template_count
    assert workbook["Value Expressions"].max_row - 1 == expression_count
    assert workbook["Warnings"].max_row - 1 == warning_count
    assert all(
        cell.data_type != "f"
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
    )
    workbook.close()


@pytest.mark.parametrize("direction", ["assembly", "parse"])
def test_committed_workbook_zip_has_no_active_or_sensitive_content(direction: str) -> None:
    path = CHAIN_ROOT / "templates" / direction / "v1" / "configuration-workbook.xlsx"
    with zipfile.ZipFile(path) as archive:
        names = archive.namelist()
        combined = b"\n".join(archive.read(name) for name in names if name.endswith((".xml", ".rels")))

    assert not any("vbaProject" in name for name in names)
    assert not any("externalLinks" in name for name in names)
    assert re.search(rb"<f(?:\s|>)", combined) is None
    assert b"<REDACTED>" not in combined
    assert b">DEBT<" not in combined
    assert b">OUR<" not in combined
