from __future__ import annotations

import json
import logging
import os
import re
import uuid
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, TypedDict

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from . import __version__
from .configuration_rules import REDACTED_VALUE, RulePackage
from .interface_standard_validator import validate_interface_standard
from .interface_template_validator import validate_interface_template
from .schemair_validator import validate_schemair


LOGGER = logging.getLogger(__name__)

WORKBOOK_FORMAT_VERSION = "v1"
SHEET_NAMES = (
    "Overview",
    "Interface Standard",
    "Interface Template",
    "Value Expressions",
    "Warnings",
    "Rule References",
    "Legend",
)
STANDARD_ACTIONS = {"CREATE", "REUSE", "UPDATE"}
EXCEL_CELL_CHARACTER_LIMIT = 32_767
ILLEGAL_CONTROL_CHARACTERS = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
READ_ONLY_FILL = PatternFill("solid", fgColor="D9E1F2")
WARNING_FILL = PatternFill("solid", fgColor="FFF2CC")
ERROR_FILL = PatternFill("solid", fgColor="F4CCCC")
INFO_FILL = PatternFill("solid", fgColor="D9EAD3")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BODY_ALIGNMENT = Alignment(vertical="top", wrap_text=True)

STANDARD_HEADERS = (
    "Field ID",
    "Sequence",
    "Field Name",
    "Field Description",
    "Parent Path",
    "Full Path",
    "Required",
    "Conditional Required",
    "Length Limit",
    "Illegal Characters",
    "XML Keys",
    "Regex",
    "Data Type",
    "SchemaIR Path",
    "Bank Required",
    "Bank Length",
    "Bank Occurs",
    "Rule Reference",
    "Difference Reason",
    "Confidence",
    "Validator Issue",
    "Execution Status",
    "Verification Status",
    "Operator Note",
)

TEMPLATE_HEADERS = (
    "Standard Field Ref",
    "Standard Role",
    "Standard Field Name",
    "Standard Parent Path",
    "Standard Full Path",
    "Standard Required",
    "Standard Length",
    "Standard Data Type",
    "Template Required",
    "Template Length",
    "Template Data Type",
    "Parse Target Ref",
    "Parse Target Name",
    "Parse Target Parent Path",
    "Parse Target Full Path",
    "Parse Target Data Type",
    "Binding Kind",
    "Value Mode",
    "Value Summary",
    "XML Key Summary",
    "Empty Handling",
    "Overlength Handling",
    "Row Limit",
    "Chinese Character Length",
    "Replacement Rule Name",
    "Rule Reference",
    "Confidence",
    "Uncertain",
    "Validator Issue",
    "Execution Status",
    "Verification Status",
    "Operator Note",
)

EXPRESSION_HEADERS = (
    "Template ID",
    "Target Field Ref",
    "Expression Scope",
    "XML Key",
    "Expression ID",
    "Parent Expression ID",
    "Sequence",
    "Mode",
    "FIELD Reference",
    "Fixed Payload Kind",
    "Fixed Value / Secure Reference",
    "Function Reference",
    "Function Parameters",
    "Mapping Rule Name",
    "Rule Reference",
)

WARNING_HEADERS = (
    "Severity",
    "Direction",
    "Standard Field Ref",
    "Category",
    "Message",
    "Rule Reference",
    "Review Disposition",
    "Omission Reason",
    "Source",
)

RULE_HEADERS = (
    "Artifact Scope",
    "Rule Package Version",
    "Rule ID",
    "Rule Title",
    "Source File / Section",
    "Used By",
)


class WorkbookIssue(TypedDict):
    code: str
    artifact: str | None
    path: str | None
    message: str


class WorkbookGenerationError(Exception):
    def __init__(self, issues: list[WorkbookIssue]) -> None:
        self.issues = tuple(issues)
        super().__init__(f"configuration workbook generation failed with {len(issues)} issue(s)")


def validate_configuration_workbook_inputs(
    *,
    schemair: dict[str, Any],
    schemair_validation_result: dict[str, Any],
    standard: dict[str, Any],
    standard_validation_result: dict[str, Any],
    template: dict[str, Any],
    template_validation_result: dict[str, Any],
    standard_rule_package: RulePackage,
    template_rule_package: RulePackage,
) -> None:
    """Fail closed unless all supplied artifacts form one exact Final chain."""

    issues: list[WorkbookIssue] = []
    _validate_rule_package(
        standard_rule_package,
        expected_version=standard.get("rulePackageVersion") if isinstance(standard, dict) else None,
        artifact="standard-rule-package",
        issues=issues,
    )
    _validate_rule_package(
        template_rule_package,
        expected_version=template.get("rulePackageVersion") if isinstance(template, dict) else None,
        artifact="template-rule-package",
        issues=issues,
    )
    if issues:
        _raise(issues)

    computed_schema_result = validate_schemair(schemair)
    _require_exact_result(
        supplied=schemair_validation_result,
        computed=computed_schema_result,
        artifact="schemair",
        issues=issues,
    )
    computed_standard_result = validate_interface_standard(
        standard,
        schemair=schemair,
        rule_package=standard_rule_package,
    )
    _require_exact_result(
        supplied=standard_validation_result,
        computed=computed_standard_result,
        artifact="standard",
        issues=issues,
    )
    computed_template_result = validate_interface_template(
        template,
        standard=standard,
        rule_package=template_rule_package,
    )
    _require_exact_result(
        supplied=template_validation_result,
        computed=computed_template_result,
        artifact="template",
        issues=issues,
    )

    identities = {
        "interfaceCode": (schemair.get("interfaceCode"), standard.get("interfaceCode"), template.get("interfaceCode")),
    }
    for name, values in identities.items():
        if len(set(values)) != 1 or not isinstance(values[0], str):
            _issue(
                issues,
                "TRUSTED_CHAIN_IDENTITY_MISMATCH",
                "trusted-chain",
                name,
                f"All artifacts must use the same {name}.",
            )
    direction = standard.get("direction")
    if direction != template.get("direction") or direction not in {"ASSEMBLY", "PARSE"}:
        _issue(
            issues,
            "TRUSTED_CHAIN_DIRECTION_MISMATCH",
            "trusted-chain",
            "direction",
            "Standard and Template must use the same supported direction.",
        )
    schema_messages = schemair.get("messages") if isinstance(schemair, dict) else None
    if not isinstance(schema_messages, list) or not any(
        isinstance(message, dict) and message.get("functionType") == direction for message in schema_messages
    ):
        _issue(
            issues,
            "TRUSTED_CHAIN_DIRECTION_MISSING",
            "schemair",
            "messages",
            "SchemaIR must contain the selected direction.",
        )
    if issues:
        _raise(issues)


def generate_configuration_workbook(
    *,
    schemair: dict[str, Any],
    schemair_validation_result: dict[str, Any],
    standard: dict[str, Any],
    standard_validation_result: dict[str, Any],
    template: dict[str, Any],
    template_validation_result: dict[str, Any],
    standard_rule_package: RulePackage,
    template_rule_package: RulePackage,
    standard_action: str,
    output_path: Path,
    generated_at: datetime,
    overwrite: bool = False,
) -> Path:
    """Generate one deterministic deliverable workbook after trusted-chain validation."""

    context = {
        "component": "configuration_workbook",
        "interface_code": schemair.get("interfaceCode") if isinstance(schemair, dict) else None,
        "direction": standard.get("direction") if isinstance(standard, dict) else None,
        "template_id": template.get("templateId") if isinstance(template, dict) else None,
    }
    LOGGER.debug("Generating configuration workbook", extra={**context, "outcome": "started"})
    try:
        validate_configuration_workbook_inputs(
            schemair=schemair,
            schemair_validation_result=schemair_validation_result,
            standard=standard,
            standard_validation_result=standard_validation_result,
            template=template,
            template_validation_result=template_validation_result,
            standard_rule_package=standard_rule_package,
            template_rule_package=template_rule_package,
        )
        _validate_generation_context(standard_action, output_path, generated_at, overwrite=overwrite)
        workbook = _build_workbook(
            schemair=schemair,
            schemair_validation_result=schemair_validation_result,
            standard=standard,
            standard_validation_result=standard_validation_result,
            template=template,
            template_validation_result=template_validation_result,
            standard_rule_package=standard_rule_package,
            template_rule_package=template_rule_package,
            standard_action=standard_action,
            generated_at=generated_at,
        )
        output = _save_workbook_atomically(workbook, Path(output_path), overwrite=overwrite)
    except WorkbookGenerationError:
        LOGGER.warning("Configuration workbook generation rejected", extra={**context, "outcome": "rejected"})
        raise
    except Exception as exc:
        LOGGER.exception("Configuration workbook generation failed", extra={**context, "outcome": "failed"})
        raise WorkbookGenerationError(
            [_workbook_issue("WORKBOOK_GENERATION_FAILED", "workbook", None, "Workbook could not be generated.")]
        ) from exc

    row_counts = {sheet.title: max(sheet.max_row - 1, 0) for sheet in workbook.worksheets}
    LOGGER.info(
        "Configuration workbook generated",
        extra={**context, "outcome": "succeeded", "sheet_count": len(workbook.sheetnames), "row_counts": row_counts},
    )
    return output


def _validate_rule_package(
    package: object,
    *,
    expected_version: object,
    artifact: str,
    issues: list[WorkbookIssue],
) -> None:
    if not isinstance(package, RulePackage):
        _issue(issues, "INVALID_RULE_PACKAGE", artifact, None, "Rule package must be a validated RulePackage.")
        return
    if package.status != "RELEASED":
        _issue(issues, "RULE_PACKAGE_NOT_RELEASED", artifact, "status", "Rule package must be RELEASED.")
    if package.version != expected_version:
        _issue(
            issues,
            "RULE_PACKAGE_VERSION_MISMATCH",
            artifact,
            "version",
            "Rule package version must match the artifact reference.",
        )


def _require_exact_result(
    *,
    supplied: object,
    computed: dict[str, Any],
    artifact: str,
    issues: list[WorkbookIssue],
) -> None:
    if supplied != computed:
        _issue(
            issues,
            "VALIDATION_RESULT_MISMATCH",
            artifact,
            "validation-result",
            "Supplied validation result must exactly match a fresh Validator result.",
        )
        return
    if computed.get("finalEligible") is not True:
        _issue(
            issues,
            "ARTIFACT_NOT_FINAL_ELIGIBLE",
            artifact,
            "finalEligible",
            "Artifact must have finalEligible=true.",
        )


def _validate_generation_context(
    standard_action: object,
    output_path: Path,
    generated_at: object,
    *,
    overwrite: object,
) -> None:
    issues: list[WorkbookIssue] = []
    if not isinstance(standard_action, str) or standard_action not in STANDARD_ACTIONS:
        _issue(
            issues,
            "INVALID_STANDARD_ACTION",
            "workbook",
            "standard_action",
            "standard_action must be exactly CREATE, REUSE, or UPDATE.",
        )
    if not isinstance(generated_at, datetime) or generated_at.tzinfo is None or generated_at.utcoffset() is None:
        _issue(
            issues,
            "INVALID_GENERATED_AT",
            "workbook",
            "generated_at",
            "generated_at must be an offset-aware datetime.",
        )
    if not isinstance(overwrite, bool):
        _issue(issues, "INVALID_OVERWRITE", "workbook", "overwrite", "overwrite must be a boolean.")
    output = Path(output_path)
    if output.suffix.lower() != ".xlsx":
        _issue(issues, "INVALID_OUTPUT_PATH", "workbook", "output_path", "Output path must use the .xlsx extension.")
    if not output.parent.exists() or not output.parent.is_dir():
        _issue(
            issues,
            "INVALID_OUTPUT_DIRECTORY",
            "workbook",
            "output_path",
            "Output parent must be an existing directory.",
        )
    if output.exists() and not overwrite:
        _issue(
            issues,
            "OUTPUT_ALREADY_EXISTS",
            "workbook",
            "output_path",
            "Output already exists; pass overwrite=True to replace it.",
        )
    if issues:
        _raise(issues)


def _build_workbook(
    *,
    schemair: dict[str, Any],
    schemair_validation_result: dict[str, Any],
    standard: dict[str, Any],
    standard_validation_result: dict[str, Any],
    template: dict[str, Any],
    template_validation_result: dict[str, Any],
    standard_rule_package: RulePackage,
    template_rule_package: RulePackage,
    standard_action: str,
    generated_at: datetime,
) -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)
    _add_overview(
        workbook,
        schemair,
        schemair_validation_result,
        standard,
        standard_validation_result,
        template,
        template_validation_result,
        standard_action,
        generated_at,
    )
    _add_standard_sheet(
        workbook,
        schemair,
        schemair_validation_result,
        standard,
        standard_validation_result,
        standard_action,
    )
    _add_template_sheet(workbook, standard, template, template_validation_result)
    _add_expression_sheet(workbook, template)
    _add_warnings_sheet(
        workbook,
        schemair_validation_result,
        standard,
        standard_validation_result,
        template,
        template_validation_result,
    )
    _add_rule_references_sheet(
        workbook,
        standard,
        template,
        standard_rule_package,
        template_rule_package,
    )
    _add_legend_sheet(workbook)
    if tuple(workbook.sheetnames) != SHEET_NAMES:
        raise WorkbookGenerationError(
            [_workbook_issue("INVALID_SHEET_ORDER", "workbook", "sheetnames", "Workbook sheet order is invalid.")]
        )
    return workbook


def _add_overview(
    workbook: Workbook,
    schemair: dict[str, Any],
    schema_result: dict[str, Any],
    standard: dict[str, Any],
    standard_result: dict[str, Any],
    template: dict[str, Any],
    template_result: dict[str, Any],
    standard_action: str,
    generated_at: datetime,
) -> None:
    message = next(item for item in schemair["messages"] if item["functionType"] == standard["direction"])
    rows = [
        ("Workbook Format Version", WORKBOOK_FORMAT_VERSION),
        ("Delivery Status", "DELIVERABLE"),
        ("Generated At", generated_at.isoformat()),
        ("Generator Version", __version__),
        ("Interface Code", schemair["interfaceCode"]),
        ("Direction", standard["direction"]),
        ("XML Encoding", message["xmlEncoding"]),
        ("SchemaIR ID", schemair["schemaId"]),
        ("SchemaIR Version", schemair["schemaVersion"]),
        ("SchemaIR Contract", schemair["contractVersion"]),
        ("SchemaIR Content Hash", schema_result["validatedArtifact"]["contentHash"]),
        ("SchemaIR Validation Status", schema_result["status"]),
        ("SchemaIR Validation Summary", _compact_json(schema_result["summary"])),
        ("Standard ID", standard["standardId"]),
        ("Standard Version", standard["standardVersion"]),
        ("Standard Contract", standard["contractVersion"]),
        ("Standard Content Hash", standard_result["validatedArtifact"]["contentHash"]),
        ("Standard Rule Package Version", standard["rulePackageVersion"]),
        ("Standard Action", standard_action),
        ("Standard Validation Status", standard_result["status"]),
        ("Standard Validation Summary", _compact_json(standard_result["summary"])),
        ("Template ID", template["templateId"]),
        ("Template Version", template["templateVersion"]),
        ("Template Contract", template["contractVersion"]),
        ("Template Content Hash", template_result["validatedArtifact"]["contentHash"]),
        ("Template Rule Package Version", template["rulePackageVersion"]),
        ("Template Validation Status", template_result["status"]),
        ("Template Validation Summary", _compact_json(template_result["summary"])),
        (
            "Update Notice",
            "需人工与目标环境现有版本对照；本工作簿不声称包含真实版本差异。"
            if standard_action == "UPDATE"
            else "",
        ),
    ]
    sheet = workbook.create_sheet("Overview")
    _write_table(sheet, ("Key", "Value"), rows)


def _add_standard_sheet(
    workbook: Workbook,
    schemair: dict[str, Any],
    schema_result: dict[str, Any],
    standard: dict[str, Any],
    standard_result: dict[str, Any],
    standard_action: str,
) -> None:
    schema_fields = _schema_fields_by_path(schemair, standard["direction"])
    standard_issues = standard_result.get("issues", [])
    schema_issues = schema_result.get("issues", [])
    if standard_action == "REUSE":
        execution_status = verification_status = "NOT_APPLICABLE"
    else:
        execution_status = "NOT_STARTED"
        verification_status = "NOT_VERIFIED"
    rows: list[tuple[Any, ...]] = []
    for field in _ordered_standard_fields(standard["fields"]):
        schema_field = schema_fields.get(field["schemaIrFieldPath"], {})
        validator_issues = [
            item for item in schema_issues if item.get("path") == field["schemaIrFieldPath"]
        ] + _issues_for_standard_field(standard_issues, field)
        rows.append(
            (
                field["fieldId"],
                field["sequence"],
                field["fieldName"],
                field["fieldDescription"],
                field["parentPath"],
                field["fullPath"],
                field["required"],
                _condition_summary(field.get("conditionalConstraints", [])),
                _constraint_summary(field["lengthLimit"]),
                _constraint_summary(field["illegalCharacters"]),
                "\n".join(item["name"] for item in field.get("xmlKeys", [])),
                _constraint_summary(field["regex"]),
                field["dataType"],
                field["schemaIrFieldPath"],
                schema_field.get("required"),
                _schema_length_summary(schema_field.get("length")),
                schema_field.get("occurs"),
                "\n".join(field.get("ruleReferences", [])),
                "\n".join(item["reason"] for item in field.get("differences", [])),
                field["confidence"],
                _issue_summary(validator_issues),
                execution_status,
                verification_status,
                "",
            )
        )
    sheet = workbook.create_sheet("Interface Standard")
    _write_table(sheet, STANDARD_HEADERS, rows)
    if rows and standard_action != "REUSE":
        _add_status_validations(sheet, STANDARD_HEADERS, len(rows))
    elif rows:
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            for header in ("Execution Status", "Verification Status", "Operator Note"):
                row[STANDARD_HEADERS.index(header)].fill = READ_ONLY_FILL


def _add_template_sheet(
    workbook: Workbook,
    standard: dict[str, Any],
    template: dict[str, Any],
    template_result: dict[str, Any],
) -> None:
    direction = template["direction"]
    standard_fields = {field["fieldId"]: field for field in standard["fields"]}
    rows: list[tuple[Any, ...]] = []
    for config in template["fieldConfigs"]:
        refs = _standard_refs_for_config(config, direction)
        referenced_fields = [standard_fields[ref] for ref in refs]
        parse_target = config.get("parseTarget") or {}
        projection = (config.get("standardTarget") or {}).get("standardProjection") or {}
        if direction == "PARSE":
            required_values = [field["required"] for field in referenced_fields]
            length_values = [_constraint_summary(field["lengthLimit"]) for field in referenced_fields]
            data_types = [field["dataType"] for field in referenced_fields]
        else:
            required_values = [projection.get("required")] if refs else []
            length_values = [_constraint_summary(projection.get("length"))] if refs else []
            data_types = [projection.get("dataType")] if refs else []
        expression = config.get("valueExpression")
        rows.append(
            (
                "\n".join(refs),
                "TARGET" if direction == "ASSEMBLY" else "SOURCE",
                _join_field_values(referenced_fields, "fieldName"),
                _join_field_values(referenced_fields, "parentPath"),
                _join_field_values(referenced_fields, "fullPath"),
                _join_field_values(referenced_fields, "required"),
                "\n".join(_constraint_summary(field["lengthLimit"]) for field in referenced_fields),
                _join_field_values(referenced_fields, "dataType"),
                "\n".join(_display(value) for value in required_values),
                "\n".join(_display(value) for value in length_values),
                "\n".join(_display(value) for value in data_types),
                parse_target.get("parseFieldRef"),
                parse_target.get("name"),
                parse_target.get("parentPath"),
                parse_target.get("fullPath"),
                parse_target.get("dataType"),
                config["bindingKind"],
                expression.get("mode") if isinstance(expression, dict) else "",
                _expression_summary(expression),
                _xml_key_summary(config.get("xmlKeyExpressions") or {}),
                config["processingPolicies"]["emptyHandling"],
                config["processingPolicies"]["overlengthHandling"],
                config["processingPolicies"]["rowLimit"],
                config["processingPolicies"]["chineseCharacterLength"],
                config["processingPolicies"]["replacementRuleName"],
                "\n".join(config.get("ruleReferences", [])),
                config["confidence"],
                config["uncertain"],
                _issue_summary(_issues_for_template_config(template_result.get("issues", []), config, refs)),
                "NOT_STARTED",
                "NOT_VERIFIED",
                "",
            )
        )
    sheet = workbook.create_sheet("Interface Template")
    _write_table(sheet, TEMPLATE_HEADERS, rows)
    if rows:
        _add_status_validations(sheet, TEMPLATE_HEADERS, len(rows))


def _add_expression_sheet(workbook: Workbook, template: dict[str, Any]) -> None:
    rows: list[tuple[Any, ...]] = []
    for config in template["fieldConfigs"]:
        target_ref = _template_target_ref(config, template["direction"])
        expression = config.get("valueExpression")
        if isinstance(expression, dict):
            rows.extend(
                _expression_rows(
                    template["templateId"],
                    target_ref,
                    "FIELD_VALUE",
                    None,
                    expression,
                )
            )
        for xml_key in sorted((config.get("xmlKeyExpressions") or {})):
            rows.extend(
                _expression_rows(
                    template["templateId"],
                    target_ref,
                    "XML_KEY",
                    xml_key,
                    config["xmlKeyExpressions"][xml_key],
                )
            )
    sheet = workbook.create_sheet("Value Expressions")
    _write_table(sheet, EXPRESSION_HEADERS, rows)


def _add_warnings_sheet(
    workbook: Workbook,
    schema_result: dict[str, Any],
    standard: dict[str, Any],
    standard_result: dict[str, Any],
    template: dict[str, Any],
    template_result: dict[str, Any],
) -> None:
    rows = _warning_rows(schema_result, standard, standard_result, template, template_result)
    sheet = workbook.create_sheet("Warnings")
    _write_table(sheet, WARNING_HEADERS, rows)
    severity_index = WARNING_HEADERS.index("Severity")
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        fill = {"ERROR": ERROR_FILL, "WARNING": WARNING_FILL, "INFO": INFO_FILL}.get(row[severity_index].value)
        if fill:
            for cell in row:
                cell.fill = fill


def _add_rule_references_sheet(
    workbook: Workbook,
    standard: dict[str, Any],
    template: dict[str, Any],
    standard_package: RulePackage,
    template_package: RulePackage,
) -> None:
    rows: set[tuple[Any, ...]] = set()
    for field in standard["fields"]:
        used_by = field["fieldId"]
        for rule_id in _all_rule_references(field):
            rule = standard_package.rules_by_id[rule_id]
            rows.add(
                (
                    "STANDARD",
                    standard_package.version,
                    rule_id,
                    rule["summary"],
                    f"rules.yaml / rules[id={rule_id}]",
                    used_by,
                )
            )
    for config in template["fieldConfigs"]:
        used_by = _template_target_ref(config, template["direction"])
        for rule_id in _all_rule_references(config):
            rule = template_package.rules_by_id[rule_id]
            rows.add(
                (
                    "TEMPLATE",
                    template_package.version,
                    rule_id,
                    rule["summary"],
                    f"rules.yaml / rules[id={rule_id}]",
                    used_by,
                )
            )
    sheet = workbook.create_sheet("Rule References")
    _write_table(sheet, RULE_HEADERS, sorted(rows, key=lambda row: (row[0], row[1], row[2], row[5])))


def _add_legend_sheet(workbook: Workbook) -> None:
    rows = [
        ("Workbook Format", WORKBOOK_FORMAT_VERSION, "固定结构版本；不是 IR contract version。"),
        ("Delivery Status", "DELIVERABLE", "全部 Final 输入和匹配校验结果通过后生成。"),
        ("Standard Action", "CREATE", "新建标准，执行状态从 NOT_STARTED 开始。"),
        ("Standard Action", "REUSE", "只读核对标准，执行与验证状态为 NOT_APPLICABLE。"),
        ("Standard Action", "UPDATE", "更新标准，但必须人工对照目标环境旧版本。"),
        ("Execution Status", "NOT_STARTED", "尚未开始配置。"),
        ("Execution Status", "IN_PROGRESS", "正在配置。"),
        ("Execution Status", "CONFIGURED", "配置已完成，可进入验证。"),
        ("Execution Status", "BLOCKED", "配置被阻塞，可恢复为 IN_PROGRESS。"),
        ("Verification Status", "NOT_VERIFIED", "尚未验证或配置变更后已重置。"),
        ("Verification Status", "PASSED", "仅 CONFIGURED 行可标记通过。"),
        ("Verification Status", "FAILED", "仅 CONFIGURED 行可标记失败。"),
        ("Binding Kind", "VALUE", "标量字段值绑定。"),
        ("Binding Kind", "STRUCTURE_ONLY", "Node/Object 结构或 XML Key 配置，不含字段值表达式。"),
        ("Binding Kind", "COLLECTION_ITEM", "重复 Standard Node 到 Parse List 元素的结构绑定。"),
        ("Value Mode", "FIXED_VALUE", "固定 LITERAL 或 SECURE_INPUT_REF；安全值本身不进入工作簿。"),
        ("Value Mode", "EMPTY", "存在配置行且明确取空值。"),
        ("Value Mode", "FIELD", "引用 FIELD catalog 或绑定 Standard source。"),
        ("Value Mode", "FUNCTION", "调用已验证的 String function。"),
        ("Value Mode", "MAPPING", "完整 String 精确映射；工作簿不复制 entries。"),
        ("Value Mode", "CONCATENATE", "按 Sequence 递归拼接子表达式。"),
        ("Blank Value Mode", "N/A", "Node/Object 的空值表示不适用，不是 UNKNOWN、omission 或 EMPTY。"),
        ("Omission", "ACCEPTED", "不存在 Template 行，但 Human Review 已接受并保留为 Warning。"),
        ("Secure Input", "SECURE_INPUT_REF", "只展示安全引用标识，不展示真实值。"),
    ]
    sheet = workbook.create_sheet("Legend")
    _write_table(sheet, ("Topic", "Value", "Meaning"), rows)


def _warning_rows(
    schema_result: dict[str, Any],
    standard: dict[str, Any],
    standard_result: dict[str, Any],
    template: dict[str, Any],
    template_result: dict[str, Any],
) -> list[tuple[Any, ...]]:
    direction = standard["direction"]
    path_to_field: dict[str, dict[str, Any]] = {}
    for field in standard["fields"]:
        path_to_field[field["schemaIrFieldPath"]] = field
        for key in field.get("xmlKeys", []):
            path_to_field[key["schemaIrFieldPath"]] = field

    conditions_by_schema_path: dict[str, tuple[dict[str, Any], list[dict[str, Any]]]] = {}
    for field in standard["fields"]:
        conditions = field.get("conditionalConstraints", [])
        if conditions:
            # SchemaIR 对同一字段只产生一条字段级提示，因此必须一次投影该字段的全部结构化条件。
            conditions_by_schema_path[field["schemaIrFieldPath"]] = (field, conditions)

    rows: list[tuple[Any, ...]] = []
    consumed_schema_issue_paths: set[tuple[str, str]] = set()
    for item in schema_result.get("issues", []):
        path = item.get("path")
        field = path_to_field.get(path)
        if field is None:
            continue
        condition_group = conditions_by_schema_path.get(path)
        if item.get("code") == "CONDITIONAL_FIELD" and condition_group:
            _, conditions = condition_group
            rows.extend(
                _condition_warning(direction, field, condition, source="SchemaIR Validator + Review")
                for condition in conditions
            )
        else:
            rows.append(
                (
                    item.get("severity"),
                    direction,
                    field["fieldId"],
                    "VALIDATOR",
                    f"{item.get('code')}: {item.get('message')}",
                    "",
                    "NOT_REQUIRED",
                    "",
                    "SchemaIR Validator",
                )
            )
        consumed_schema_issue_paths.add((str(item.get("code")), str(path)))

    for field in standard["fields"]:
        for condition in field.get("conditionalConstraints", []):
            key = ("CONDITIONAL_FIELD", field["schemaIrFieldPath"])
            if key not in consumed_schema_issue_paths:
                rows.append(_condition_warning(direction, field, condition, source="Standard + Review"))
        for difference in field.get("differences", []):
            rows.append(
                (
                    "WARNING",
                    direction,
                    field["fieldId"],
                    "SCHEMA_STANDARD_DIFFERENCE",
                    f"{difference['property']}: {difference['reason']}",
                    "\n".join(difference.get("ruleReferences", [])),
                    "ACCEPTED",
                    "",
                    "InterfaceStandardIR + Review",
                )
            )

    consumed_template_issue_paths: set[str] = set()
    for omission in template.get("omissions", []):
        field_ref = omission["standardFieldRef"]
        issue_path = f"omissions.{field_ref}"
        matching = next(
            (item for item in template_result.get("issues", []) if item.get("path") == issue_path),
            None,
        )
        consumed_template_issue_paths.add(issue_path)
        message = matching.get("message") if matching else "ASSEMBLY Standard field is intentionally omitted."
        rows.append(
            (
                "WARNING",
                direction,
                field_ref,
                "MISSING_TEMPLATE_FIELD",
                f"MISSING_TEMPLATE_FIELD: {message}",
                "",
                omission["reviewDisposition"],
                omission["reason"],
                "Template Validator + Review",
            )
        )

    for source, result in (
        ("InterfaceStandardIR Validator", standard_result),
        ("InterfaceTemplateIR Validator", template_result),
    ):
        for item in result.get("issues", []):
            if item.get("path") in consumed_template_issue_paths:
                continue
            rows.append(
                (
                    item.get("severity"),
                    direction,
                    _field_ref_from_issue(item, standard),
                    "VALIDATOR",
                    f"{item.get('code')}: {item.get('message')}",
                    "",
                    "NOT_REQUIRED",
                    "",
                    source,
                )
            )

    severity_order = {"ERROR": 0, "WARNING": 1, "INFO": 2}
    return sorted(rows, key=lambda row: (severity_order.get(str(row[0]), 9), str(row[3]), str(row[2]), str(row[8])))


def _condition_warning(
    direction: str,
    field: dict[str, Any],
    condition: dict[str, Any],
    *,
    source: str,
) -> tuple[Any, ...]:
    operator = condition["operator"]
    operand = f" {condition['literal']}" if operator == "EQUALS" else ""
    message = (
        f"{condition['controllingFieldRef']} {operator}{operand} => "
        f"{condition['targetFieldRef']} {condition['effect']}; {condition['sourceText']}"
    )
    return (
        "INFO",
        direction,
        field["fieldId"],
        "BANK_CONDITIONAL_CONSTRAINT",
        message,
        "\n".join(condition.get("ruleReferences", [])),
        "ACCEPTED",
        "",
        source,
    )


def _expression_rows(
    template_id: str,
    target_ref: str,
    scope: str,
    xml_key: str | None,
    root: dict[str, Any],
) -> list[tuple[Any, ...]]:
    rows: list[tuple[Any, ...]] = []
    prefix = f"{scope.lower()}:{target_ref}:{xml_key or 'field'}"

    def visit(expression: dict[str, Any], index_path: tuple[int, ...], parent_id: str) -> None:
        expression_id = f"{prefix}:{'.'.join(str(index) for index in index_path)}"
        payload = expression.get("payload") or {}
        field_reference = expression.get("standardFieldRef") or expression.get("assemblyFieldRef")
        arguments = sorted(expression.get("arguments") or [], key=lambda item: item["position"])
        normalized_arguments = [
            {
                "position": item["position"],
                "kind": item["kind"],
                "valueOrRef": item.get("value") or item.get("standardFieldRef") or item.get("assemblyFieldRef"),
            }
            for item in arguments
        ]
        rows.append(
            (
                template_id,
                target_ref,
                scope,
                xml_key or "",
                expression_id,
                parent_id,
                expression["sequence"],
                expression["mode"],
                field_reference,
                payload.get("kind"),
                payload.get("value"),
                expression.get("functionCode"),
                _compact_json(normalized_arguments) if normalized_arguments else "",
                expression.get("mappingRuleName"),
                "\n".join(expression.get("ruleReferences", [])),
            )
        )
        children = sorted(expression.get("children") or [], key=lambda item: item["sequence"])
        for child_index, child in enumerate(children):
            visit(child, (*index_path, child_index), expression_id)

    visit(root, (0,), "")
    return rows


def _save_workbook_atomically(workbook: Workbook, output_path: Path, *, overwrite: bool) -> Path:
    temporary = output_path.parent / f".{output_path.name}.{uuid.uuid4().hex}.tmp.xlsx"
    try:
        workbook.save(temporary)
        loaded = load_workbook(temporary, data_only=False, read_only=False, keep_links=False)
        try:
            if tuple(loaded.sheetnames) != SHEET_NAMES:
                raise WorkbookGenerationError(
                    [_workbook_issue("WORKBOOK_READBACK_FAILED", "workbook", "sheetnames", "Saved workbook sheet order changed.")]
                )
        finally:
            loaded.close()
        if overwrite:
            os.replace(temporary, output_path)
        else:
            # 同目录 hard link 提供“不覆盖”的原子发布；目标在竞争窗口出现时创建失败。
            try:
                os.link(temporary, output_path)
            except FileExistsError as exc:
                raise WorkbookGenerationError(
                    [_workbook_issue("OUTPUT_ALREADY_EXISTS", "workbook", "output_path", "Output already exists.")]
                ) from exc
            temporary.unlink()
    except OSError as exc:
        raise WorkbookGenerationError(
            [_workbook_issue("WORKBOOK_WRITE_FAILED", "workbook", "output_path", "Workbook could not be saved atomically.")]
        ) from exc
    finally:
        if temporary.exists():
            temporary.unlink()
    return output_path


def _write_table(sheet: Any, headers: Iterable[str], rows: Iterable[Iterable[Any]]) -> None:
    header_values = tuple(headers)
    _write_row(sheet, 1, header_values)
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = BODY_ALIGNMENT
    for row_index, row in enumerate(rows, start=2):
        _write_row(sheet, row_index, row)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(header_values))}{max(sheet.max_row, 1)}"
    sheet.sheet_view.showGridLines = False
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            cell.alignment = BODY_ALIGNMENT
    for index, header in enumerate(header_values, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = _column_width(header)


def _write_row(sheet: Any, row_index: int, values: Iterable[Any]) -> None:
    for column_index, value in enumerate(values, start=1):
        _set_cell(sheet.cell(row=row_index, column=column_index), value)


def _set_cell(cell: Cell, value: Any) -> None:
    if value is None:
        cell.value = None
        return
    if isinstance(value, str):
        if REDACTED_VALUE in value:
            raise WorkbookGenerationError(
                [_workbook_issue("REDACTED_VALUE_REJECTED", "workbook", cell.coordinate, "Redacted placeholder cannot be written.")]
            )
        if ILLEGAL_CONTROL_CHARACTERS.search(value):
            raise WorkbookGenerationError(
                [_workbook_issue("ILLEGAL_CELL_CHARACTER", "workbook", cell.coordinate, "Cell contains an illegal control character.")]
            )
        if len(value) > EXCEL_CELL_CHARACTER_LIMIT:
            raise WorkbookGenerationError(
                [_workbook_issue("CELL_VALUE_TOO_LONG", "workbook", cell.coordinate, "Cell exceeds Excel's character limit.")]
            )
        cell.value = value
        # openpyxl 会把等号开头的普通字符串推断为公式；可信交付物必须始终保存来源文本。
        cell.data_type = "s"
        cell.number_format = "@"
        return
    if isinstance(value, (bool, int, float)):
        cell.value = value
        return
    _set_cell(cell, _display(value))


def _add_status_validations(sheet: Any, headers: tuple[str, ...], row_count: int) -> None:
    execution_column = get_column_letter(headers.index("Execution Status") + 1)
    verification_column = get_column_letter(headers.index("Verification Status") + 1)
    execution = DataValidation(
        type="list",
        formula1='"NOT_STARTED,IN_PROGRESS,CONFIGURED,BLOCKED"',
        allow_blank=False,
        showErrorMessage=True,
        errorStyle="stop",
        errorTitle="Invalid execution status",
        error="Choose one supported execution status.",
    )
    verification = DataValidation(
        type="list",
        formula1='"NOT_VERIFIED,PASSED,FAILED"',
        allow_blank=False,
        showErrorMessage=True,
        errorStyle="stop",
        errorTitle="Invalid verification status",
        error="Choose one supported verification status.",
    )
    sheet.add_data_validation(execution)
    sheet.add_data_validation(verification)
    execution.add(f"{execution_column}2:{execution_column}{row_count + 1}")
    verification.add(f"{verification_column}2:{verification_column}{row_count + 1}")


def _schema_fields_by_path(schemair: dict[str, Any], direction: str) -> dict[str, dict[str, Any]]:
    result = {field["path"]: field for field in schemair["envelope"]["fields"]}
    message = next(item for item in schemair["messages"] if item["functionType"] == direction)
    result.update({field["path"]: field for field in message["fields"]})
    return result


def _ordered_standard_fields(fields: list[dict[str, Any]]) -> list[dict[str, Any]]:
    children: defaultdict[str, list[dict[str, Any]]] = defaultdict(list)
    for field in fields:
        children[field["parentPath"]].append(field)
    for values in children.values():
        values.sort(key=lambda item: (item["sequence"], item["fieldId"]))
    ordered: list[dict[str, Any]] = []

    def visit(parent_path: str) -> None:
        for field in children.get(parent_path, []):
            ordered.append(field)
            visit(field["fullPath"])

    visit("Root")
    if len(ordered) != len(fields):
        raise WorkbookGenerationError(
            [_workbook_issue("STANDARD_TREE_INCOMPLETE", "standard", "fields", "Standard field tree cannot be fully traversed.")]
        )
    return ordered


def _standard_refs_for_config(config: dict[str, Any], direction: str) -> list[str]:
    if direction == "ASSEMBLY":
        target = config.get("standardTarget") or {}
        return [target["standardFieldRef"]] if target.get("standardFieldRef") else []
    refs: list[str] = []
    source = config.get("standardSource") or {}
    if source.get("standardFieldRef"):
        refs.append(source["standardFieldRef"])

    def visit(expression: object) -> None:
        if not isinstance(expression, dict):
            return
        reference = expression.get("standardFieldRef")
        if isinstance(reference, str) and reference not in refs:
            refs.append(reference)
        for argument in sorted(expression.get("arguments") or [], key=lambda item: item["position"]):
            reference = argument.get("standardFieldRef")
            if isinstance(reference, str) and reference not in refs:
                refs.append(reference)
        for child in sorted(expression.get("children") or [], key=lambda item: item["sequence"]):
            visit(child)

    visit(config.get("valueExpression"))
    return refs


def _template_target_ref(config: dict[str, Any], direction: str) -> str:
    if direction == "ASSEMBLY":
        return config["standardTarget"]["standardFieldRef"]
    return config["parseTarget"]["parseFieldRef"]


def _all_rule_references(value: object) -> set[str]:
    result: set[str] = set()
    if isinstance(value, dict):
        references = value.get("ruleReferences")
        if isinstance(references, list):
            result.update(item for item in references if isinstance(item, str))
        for child in value.values():
            result.update(_all_rule_references(child))
    elif isinstance(value, list):
        for child in value:
            result.update(_all_rule_references(child))
    return result


def _issues_for_standard_field(issues: list[dict[str, Any]], field: dict[str, Any]) -> list[dict[str, Any]]:
    field_id = field["fieldId"]
    full_path = field["fullPath"]
    return [item for item in issues if field_id in str(item.get("path")) or full_path in str(item.get("path"))]


def _issues_for_template_config(
    issues: list[dict[str, Any]],
    config: dict[str, Any],
    standard_refs: list[str],
) -> list[dict[str, Any]]:
    target = (config.get("parseTarget") or {}).get("parseFieldRef")
    tokens = [*standard_refs, target]
    return [item for item in issues if any(token and token in str(item.get("path")) for token in tokens)]


def _field_ref_from_issue(issue: dict[str, Any], standard: dict[str, Any]) -> str:
    path = str(issue.get("path") or "")
    for field in standard["fields"]:
        if field["fieldId"] in path or field["fullPath"] in path:
            return field["fieldId"]
    return ""


def _condition_summary(conditions: list[dict[str, Any]]) -> str:
    values = []
    for condition in conditions:
        operand = f" {condition['literal']}" if condition["operator"] == "EQUALS" else ""
        values.append(
            f"{condition['controllingFieldRef']} {condition['operator']}{operand} => {condition['effect']}"
        )
    return "\n".join(values)


def _constraint_summary(value: object) -> str:
    if not isinstance(value, dict):
        return ""
    state = value.get("state")
    if state != "VALUE":
        return _display(state)
    if "min" in value or "max" in value:
        summary = f"{_display(value.get('min'))}..{_display(value.get('max'))}"
        details = [
            f"{name}={value[name]}"
            for name in ("precision", "scale")
            if value.get(name) is not None
        ]
        return "; ".join([summary, *details])
    return _display(value.get("value"))


def _schema_length_summary(value: object) -> str:
    if not isinstance(value, dict):
        return ""
    if value.get("raw"):
        return str(value["raw"])
    if value.get("min") is None and value.get("max") is None:
        return "NO_CONSTRAINT"
    return f"{_display(value.get('min'))}..{_display(value.get('max'))}"


def _expression_summary(expression: object) -> str:
    if not isinstance(expression, dict):
        return ""
    mode = expression["mode"]
    if mode == "FIXED_VALUE":
        return f"FIXED_VALUE({expression['payload']['kind']})"
    if mode == "FIELD":
        reference = expression.get("standardFieldRef") or expression.get("assemblyFieldRef")
        return f"FIELD({reference})"
    if mode == "FUNCTION":
        return f"FUNCTION({expression['functionCode']})"
    if mode == "MAPPING":
        return f"MAPPING({expression['mappingRuleName']})"
    if mode == "CONCATENATE":
        return f"CONCATENATE({len(expression.get('children', []))})"
    return mode


def _xml_key_summary(expressions: dict[str, dict[str, Any]]) -> str:
    return "\n".join(f"{key}={_expression_summary(expressions[key])}" for key in sorted(expressions))


def _issue_summary(issues: Iterable[dict[str, Any]]) -> str:
    return "\n".join(f"{item.get('severity')} {item.get('code')}: {item.get('message')}" for item in issues)


def _join_field_values(fields: list[dict[str, Any]], name: str) -> str:
    return "\n".join(_display(field.get(name)) for field in fields)


def _compact_json(value: object) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"), allow_nan=False)


def _display(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (dict, list)):
        return _compact_json(value)
    return str(value)


def _column_width(header: str) -> int:
    lowered = header.lower()
    if any(token in lowered for token in ("description", "message", "evidence", "reason", "summary", "note")):
        return 60
    if any(token in lowered for token in (" id", " ref", "path", "hash", "expression id")) or header.endswith("ID"):
        return 42
    if any(token in lowered for token in ("sequence", "count", "required", "uncertain", "row limit", "confidence")):
        return 14
    if any(token in lowered for token in ("status", "action", "mode", "kind", "type", "severity", "direction", "version")):
        return 20
    return 28


def _workbook_issue(code: str, artifact: str | None, path: str | None, message: str) -> WorkbookIssue:
    return {"code": code, "artifact": artifact, "path": path, "message": message}


def _issue(
    issues: list[WorkbookIssue],
    code: str,
    artifact: str | None,
    path: str | None,
    message: str,
) -> None:
    issues.append(_workbook_issue(code, artifact, path, message))


def _raise(issues: list[WorkbookIssue]) -> None:
    raise WorkbookGenerationError(issues)
